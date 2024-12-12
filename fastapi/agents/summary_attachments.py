import os
import csv
import base64
import logging
from typing import Dict, Optional
import fitz
from docx import Document
import mammoth
from openpyxl import load_workbook
from langchain_openai import ChatOpenAI
from langchain.schema import HumanMessage
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Initialize logging
logger = logging.getLogger(__name__)

def encode_image_to_base64(logger, image_path):
    logger.info(f"Ariflow - encode_image_to_base64 - Encoding image to base64")
    try:
        with open(image_path, 'rb') as img_file:
            img_base64 = base64.b64encode(img_file.read()).decode('utf-8')
        logger.info(f"Ariflow - encode_image_to_base64 - Image encoded to base64 successfully")
        return img_base64
    except Exception as e:
        logger.error(f"Ariflow - encode_image_to_base64 - Error encoding image to base64: {e}")
        return None
 
def image_summarize(logger, img_base64, prompt):
    logger.info(f"Ariflow - image_summarize - Summarizing image with GPT")
    try:
        chat = ChatOpenAI(
            model="gpt-4o", 
            max_tokens=1024,
            api_key=os.getenv("OPENAI_API_KEY")
        )

        msg = chat.invoke(
            [
                HumanMessage(
                    content=[
                        {
                            "type": "text", 
                            "text": prompt
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"},
                        },
                    ]
                )
            ]
        )
        logger.info(f"Ariflow - image_summarize - Summary generated successfully")
        return msg.content
    
    except Exception as e:
        logger.error(f"Ariflow - image_summarize - Error generating summary with GPT-4o: {e}")
        return None

def parse_images(logger, image_path):
    logger.info(f"Ariflow - parse_images - Generating summaries for images in {image_path}")
    prompt = (
        "You are an assistant tasked with summarizing images for retrieval via RAGs. "
        "These summaries will be embedded and used to retrieve the raw image via RAGs. "
        "Give a concise summary of the image that is well optimized for retrieval via RAGs."
    )
    
    if os.path.isfile(image_path):
        logger.info(f"Ariflow - parse_images - Processing image")

        # Encode image to base64
        image_base64 = encode_image_to_base64(logger, image_path)
        if image_base64:
            image_summary = image_summarize(logger, image_base64, prompt)
            logger.info(f"Ariflow - parse_images - Image {image_path} summary: {image_summary}")
        
            if image_summary:
                return image_summary
            else:
                return "Failed to summarize image"
        else:
            return f"Failed to encode image {image_path}"

def parse_pdf_files(logger, file_path: str, page_limit: int = 5) -> str:
    """Process PDF files with page limit."""
    try:
        pdf_document = fitz.open(file_path)
        content = ""
        pages_to_process = min(page_limit, len(pdf_document))
        
        for page_num in range(pages_to_process):
            page = pdf_document[page_num]
            content += f"\n--- Page {page_num + 1} ---\n"
            content += page.get_text()
        
        if len(pdf_document) > pages_to_process:
            content += f"\n... ({len(pdf_document) - page_limit} remaining pages truncated) ..."
            
        pdf_document.close()
        return content.strip()
    except Exception as e:
        logger.error(f"Error parsing PDF file {file_path}: {str(e)}")
        return f"Error parsing PDF file {file_path}: {str(e)}"

def parse_excel_files(logger, file_path: str) -> str:
    """Process Excel files with row limit."""
    try:
        workbook = load_workbook(file_path, data_only=True)
        content = ""
        sheets_processed = 0
        rows_per_sheet = 100
        
        for sheet_name in workbook.sheetnames:
            if sheets_processed >= 5:  # Limit to first 5 sheets
                content += f"\n... (remaining {len(workbook.sheetnames) - sheets_processed} sheets truncated) ..."
                break
                
            sheet = workbook[sheet_name]
            content += f"\nSheet: {sheet_name}\n"
            
            # Process only first 100 rows of each sheet
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                if row_count >= rows_per_sheet:
                    content += "\n... (remaining rows truncated) ...\n"
                    break
                    
                content += ", ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
                row_count += 1
                
            sheets_processed += 1
            
        return content.strip()
    except Exception as e:
        logger.error(f"Error parsing XLSX file {file_path}: {str(e)}")
        return f"Error parsing XLSX file {file_path}: {str(e)}"

def parse_word_file(logger, file_path: str) -> str:
    """Process Word documents."""
    try:
        file_extension = os.path.splitext(file_path)[-1].lower()
        content = ""

        if file_extension == ".docx":
            doc = Document(file_path)
            paragraphs = doc.paragraphs
            
            # Calculate number of paragraphs per page (assuming ~40 paragraphs per page)
            paragraphs_per_page = 40
            page_limit = 5 * paragraphs_per_page
            
            logger.info(f"Processing first {min(len(paragraphs), page_limit)} paragraphs of Word file: {file_path}")
            
            # Get first 5 pages worth of paragraphs
            content = "\n".join([para.text for para in paragraphs[:page_limit]])
            
            if len(paragraphs) > page_limit:
                content += f"\n... ({len(paragraphs) - page_limit} remaining paragraphs truncated) ..."
            
        elif file_extension == ".doc":
            with open(file_path, "rb") as doc_file:
                result = mammoth.extract_raw_text(doc_file)
                full_text = result.value
                
                # Rough approximation of 5 pages (assuming 3000 characters per page)
                chars_per_page = 3000
                page_limit = 5 * chars_per_page
                
                if len(full_text) > page_limit:
                    content = full_text[:page_limit] + "\n... (remaining content truncated) ..."
                else:
                    content = full_text
        else:
            content = f"Unsupported file type: {file_extension}"
            
        return content
    except Exception as e:
        logger.error(f"Error parsing file {file_path}: {str(e)}")
        return f"Error parsing file {file_path}: {str(e)}"

def parse_txt_files(logger, file_path: str, char_limit: int = 10000) -> str:
    """Process text files with character limit."""
    try:
        with open(file_path, "r", encoding="utf-8") as txt_file:
            content = txt_file.read(char_limit)
            if len(content) >= char_limit:
                content += "\n... (remaining content truncated) ..."
        return content
    except Exception as e:
        logger.error(f"Error processing TXT {file_path}: {str(e)}")
        return f"Error processing TXT: {str(e)}"

def parse_csv_files(logger, file_path: str, row_limit: int = 100) -> str:
    """Process CSV files with row limit."""
    try:
        content = []
        with open(file_path, 'r') as file:
            csv_reader = csv.reader(file)
            header = next(csv_reader, None)
            if header:
                content.append(",".join(header))
            
            for idx, row in enumerate(csv_reader):
                if idx >= row_limit:
                    content.append("... (remaining rows truncated) ...")
                    break
                content.append(",".join(row))
        
        return "\n".join(content)
    except Exception as e:
        logger.error(f"Error processing CSV {file_path}: {str(e)}")
        return f"Error processing CSV: {str(e)}"