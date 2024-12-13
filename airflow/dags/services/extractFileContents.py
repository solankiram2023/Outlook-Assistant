import os
import csv
import json
import base64
import openai
from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from langchain.schema import HumanMessage

from docx import Document
import mammoth
from openpyxl import load_workbook
import fitz

# Loading environment variables
load_dotenv()

# Sub function to convert images to base64
def encode_image_to_base64(logger, image_path):
    logger.info(f"Ariflow - encode_image_to_base64 - Encoding image to base64")
    try:
        with open(image_path, 'rb') as img_file:
            img_base64 = base64.b64encode(img_file.read()).decode('utf-8')
        logger.info(f"Ariflow - encode_image_to_base64 - Image encoded to base64 successfully")
        return img_base64
    except Exception as e:
        logger.error(f"Ariflow - encode_image_to_base64 - Error encoding image to base64: {e}")
        return None
 
# Sub function to summarize images using OpenAI  
def image_summarize(logger, img_base64, prompt):
    logger.info(f"Ariflow - image_summarize - Summarizing image with GPT")
    try:

        chat = ChatOpenAI(
            model       = "gpt-4o", 
            max_tokens  = 1024,
            api_key     = os.getenv("OPENAI_API_KEY")
        )

        msg = chat.invoke(
            [
                HumanMessage(
                    content=[
                        {
                            "type": "text", 
                            "text": prompt
                        },
                        {
                            "type"      : "image_url",
                            "image_url" : {"url": f"data:image/jpeg;base64,{img_base64}"},
                        },
                    ]
                )
            ]
        )
        logger.info(f"Ariflow - image_summarize - Summary generated successfully")
        return msg.content
    
    except Exception as e:
            logger.error(f"Ariflow - image_summarize - Error generating summary with GPT-4o: {e}")
            return None

# Function to parse images and extract contents
def parse_images(logger, image_path):
    logger.info(f"Ariflow - parse_images - Generating summaries for images in {image_path}")
    prompt = (
        "You are an assistant tasked with summarizing images for retrieval via RAGs. "
        "These summaries will be embedded and used to retrieve the raw image via RAGs. "
        "Give a concise summary of the image that is well optimized for retrieval via RAGs."
    )
    
    if os.path.isfile(image_path):
        logger.info(f"Ariflow - parse_images - Processing image")

        # Encode image to base64
        image_base64 = encode_image_to_base64(logger, image_path)
        if image_base64:
            image_summary = image_summarize(logger, image_base64, prompt)
            logger.info(f"Ariflow - parse_images - Image {image_path} summary: {image_summary}")
        
            if image_summary:
                return image_summary
            else:
                return "Failed to summarize image"
        else:
            return f"Failed to encode image {image_path}"

# Function to parse CSV files and extract contents
def parse_csv_files(logger, csv_file_path):
    logger.info(f"Ariflow - parse_csv_files - Extarcting contents from csv file: {csv_file_path}")

    extracted_contents = ""

    try:
        # Open the CSV file for reading
        with open(csv_file_path, 'r') as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                # Join the row contents and append to the extracted contents
                extracted_contents += ", ".join(row) + "\n"
    except Exception as e:
        logger.error(f"Airflow - parse_csv_files - Error processing CSV file: {e}")
        extracted_contents = f"Error processing CSV file {csv_file_path}: {str(e)}"

    return extracted_contents

# Parsing Word Document files
def parse_word_file(logger, file_path):
    try:
        file_extension = os.path.splitext(file_path)[-1].lower()

        if file_extension == ".docx":
            # Use python-docx for .docx files
            doc = Document(file_path)
            content = "\n".join([para.text for para in doc.paragraphs])
        elif file_extension == ".doc":
            # Use mammoth for .doc files
            with open(file_path, "rb") as doc_file:
                result = mammoth.extract_raw_text(doc_file)
                content = result.value  # Extracted text
        else:
            content = f"Unsupported file type: {file_extension}"
    except Exception as e:
        content = f"Error parsing file {file_path}: {str(e)}"

    return content

# Parsing txt files
def parse_txt_files(logger, file_path):
    try:
        with open(file_path, "r", encoding="utf-8") as txt_file:
            content = txt_file.read()
        return content
    except Exception as e:
        return f"Error parsing file {file_path}: {str(e)}"
    

# Parsing Spreadsheets
def parse_excel_files(logger, file_path):
    try:
        workbook = load_workbook(file_path, data_only=True)
        content = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                content += ", ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
        return content.strip()
    except Exception as e:
        return f"Error parsing XLSX file {file_path}: {str(e)}"


def parse_pdf_files(logger, file_path):
    try:
        pdf_document = fitz.open(file_path)
        content = ""
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            content += page.get_text()
        pdf_document.close()
        return content.strip()
    except Exception as e:
        return f"Error parsing PDF file {file_path}: {str(e)}"